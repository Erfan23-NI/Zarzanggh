from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup,
    KeyboardButton, ReplyKeyboardMarkup
)
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, filters, ContextTypes
)
import logging
from datetime import datetime
import os
import sqlite3
import uuid
import threading
import time
import requests
from flask import Flask, Response
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# تنظیمات لاگ‌گیری
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# تنظیمات اصلی
TOKEN = os.environ.get('TOKEN', 'YOUR_BOT_TOKEN')
ADMIN_CHAT_ID = int(os.environ.get('ADMIN_CHAT_ID', 86101721))
DB_NAME = "bot_data.db"
PORT = int(os.environ.get('PORT', 10000))

# سیستم بیدار ماندن
PING_URL = f"https://{os.environ.get('RENDER_EXTERNAL_HOSTNAME', 'your-app-name.onrender.com')}"
PING_INTERVAL = 300

def keep_alive():
    while True:
        try:
            logger.info(f"🔄 ارسال پینگ به {PING_URL}...")
            response = requests.get(PING_URL)
            logger.info(f"✅ پاسخ پینگ: {response.status_code}")
        except Exception as e:
            logger.error(f"❌ خطا در ارسال پینگ: {e}")
        time.sleep(PING_INTERVAL)

def start_keep_alive():
    thread = threading.Thread(target=keep_alive, daemon=True)
    thread.start()
    logger.info("📡 فعال‌سازی سیستم بیدار ماندن ربات...")

# سرور وب
app = Flask(__name__)

@app.route('/')
def home():
    return "✅ ربات زرزنگ با موفقیت در حال اجراست!"

@app.route('/health')
def health_check():
    return Response(status=200)

def run_web_server():
    app.run(host='0.0.0.0', port=PORT)

# --- توابع دیتابیس ---
def init_db():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute('''
    CREATE TABLE IF NOT EXISTS verified_users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER UNIQUE,
        phone TEXT,
        full_name TEXT,
        nid TEXT,
        registration_date TEXT
    )
    ''')
    cur.execute('''
    CREATE TABLE IF NOT EXISTS pending_verification (
        id TEXT PRIMARY KEY,
        user_id INTEGER,
        phone TEXT,
        full_name TEXT,
        nid TEXT,
        file_id TEXT,
        date TEXT
    )
    ''')
    conn.commit()
    conn.close()

def save_verified_user(user_data):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute('''
    INSERT OR REPLACE INTO verified_users
    (user_id, phone, full_name, nid, registration_date)
    VALUES (?, ?, ?, ?, ?)
    ''', (
        user_data['user_id'],
        user_data['phone'],
        user_data['full_name'],
        user_data['nid'],
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ))
    conn.commit()
    conn.close()

def save_pending_verification(user_data):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute('''
    INSERT INTO pending_verification
    (id, user_id, phone, full_name, nid, file_id, date)
    VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (
        str(uuid.uuid4()),
        user_data['user_id'],
        user_data['phone'],
        user_data['full_name'],
        user_data['nid'],
        user_data['file_id'],
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ))
    conn.commit()
    conn.close()

def get_pending_verification(user_id):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute('SELECT * FROM pending_verification WHERE user_id = ?', (user_id,))
    result = cur.fetchone()
    conn.close()
    return result

def remove_pending_verification(user_id):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute('DELETE FROM pending_verification WHERE user_id = ?', (user_id,))
    conn.commit()
    conn.close()

def create_excel_file():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM verified_users")
    users = cursor.fetchall()
    conn.close()

    if not users:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "کاربران تایید شده"

    headers = ["ردیف", "آیدی کاربر", "تلفن", "نام کامل", "کد ملی", "تاریخ ثبت‌نام"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for idx, user in enumerate(users, start=1):
        ws.append([
            idx,
            user[1],
            user[2],
            user[3],
            user[4],
            user[5]
        ])

    column_widths = [8, 15, 15, 25, 15, 20]
    for i, width in enumerate(column_widths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"users_{timestamp}.xlsx"
    wb.save(filename)

    return filename

# --- کیبوردها ---
def start_keyboard():
    keyboard = [
        [InlineKeyboardButton("تلگرام زرزنگ", url="https://t.me/ddingooa"),
         InlineKeyboardButton("یوتیوب زرزنگ", url="https://www.youtube.com/zerzang")],
        [InlineKeyboardButton("اینستاگرام زرزنگ", url="https://instagram.com/zerzang"),
         InlineKeyboardButton("سایت زرزنگ", url="https://zerzang.com")],
        [InlineKeyboardButton("ثبت نام سیگنال فیوچرز", callback_data="register_signal")],
        [InlineKeyboardButton("پشتیبانی", callback_data="support")],
        [InlineKeyboardButton("محاسبه حد ضرر", callback_data="calc_stop_loss")],
        [InlineKeyboardButton("صرافی‌ها و بروکرهای ویژه", callback_data="exchanges_brokers")]
    ]
    return InlineKeyboardMarkup(keyboard)

def back_button():
    return InlineKeyboardMarkup([[InlineKeyboardButton("🔙 بازگشت به منو", callback_data="back_to_menu")]])

def phone_request_keyboard():
    return ReplyKeyboardMarkup(
        [
            [KeyboardButton("📱 ارسال شماره", request_contact=True)],
            [KeyboardButton("🔙 بازگشت به منو")]
        ],
        resize_keyboard=True,
        one_time_keyboard=True
    )

# --- سیستم سیگنال‌دهی ---
def create_signal_keyboard(leverage):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("۰.۵٪ سرمایه", callback_data=f"capital_percent:0.5:{leverage}")],
        [InlineKeyboardButton("۱٪ سرمایه", callback_data=f"capital_percent:1:{leverage}")],
        [InlineKeyboardButton("۱.۵٪ سرمایه", callback_data=f"capital_percent:1.5:{leverage}")],
        [InlineKeyboardButton("۲٪ سرمایه", callback_data=f"capital_percent:2:{leverage}")],
        [InlineKeyboardButton("۳٪ سرمایه", callback_data=f"capital_percent:3:{leverage}")],
        [InlineKeyboardButton("۴٪ سرمایه", callback_data=f"capital_percent:4:{leverage}")],
        [InlineKeyboardButton("۵٪ سرمایه", callback_data=f"capital_percent:5:{leverage}")]
    ])

def format_signal_message(entry, sl, tp, leverage):
    # تشخیص نوع پوزیشن
    position_type = "Long" if float(sl) < float(entry) else "Short"
    
    # محاسبه درصد ضرر
    loss_percent = abs((float(sl) - float(entry)) / float(entry)) * 100
    
    return (
        f"📈 <b>سیگنال معاملاتی جدید</b>\n\n"
        f"📍 نوع پوزیشن: <b>{position_type}</b>\n"
        f"🎯 قیمت ورود: <b>{entry}</b>\n"
        f"🛑 حد ضرر (SL): <b>{sl}</b>\n"
        f"✅ حد سود (TP): <b>{tp}</b>\n"
        f"⚖️ اهرم (Leverage): <b>{leverage}x</b>\n"
        f"📉 درصد ضرر: <b>{loss_percent:.2f}%</b>\n\n"
        f"<i>لطفاً درصدی از سرمایه که مایلید درگیر این معامله شود را انتخاب کنید:</i>"
    )

async def send_signal_to_users(context: ContextTypes.DEFAULT_TYPE, entry, sl, tp, leverage):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute('SELECT user_id FROM verified_users')
    verified_users = [row[0] for row in cur.fetchall()]
    conn.close()

    success = 0
    failed = 0
    message = format_signal_message(entry, sl, tp, leverage)
    keyboard = create_signal_keyboard(leverage)

    for user_id in verified_users:
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text=message,
                parse_mode='HTML',
                reply_markup=keyboard
            )
            success += 1
        except Exception as e:
            logger.error(f"Error sending signal to {user_id}: {e}")
            failed += 1

    return success, failed, len(verified_users)

def calculate_trade_amount(capital, percent, leverage):
    """محاسبه سرمایه ورودی به معامله با اعمال اهرم"""
    return capital * (percent / 100) * leverage

def calculate_loss_amount(trade_amount, entry, sl):
    """محاسبه ضرر دلاری در صورت فعال شدن SL"""
    return trade_amount * (abs(float(entry) - float(sl)) / float(entry))

# --- دستورات ربات ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text(
        "سلام! به ربات زرزنگ خوش آمدید.\nلطفا گزینه مورد نظر را انتخاب کنید:",
        reply_markup=start_keyboard()
    )

async def admin_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.from_user.id != ADMIN_CHAT_ID:
        await update.message.reply_text("❌ شما دسترسی ندارید!")
        return

    if context.user_data.get('admin_broadcast_mode'):
        conn = sqlite3.connect(DB_NAME)
        cur = conn.cursor()
        cur.execute('SELECT user_id FROM verified_users')
        verified_users = [row[0] for row in cur.fetchall()]
        conn.close()

        message = update.message.text
        success = 0
        failed = 0

        for user_id in verified_users:
            try:
                await context.bot.send_message(
                    chat_id=user_id,
                    text=message
                )
                success += 1
            except Exception as e:
                logger.error(f"Error sending to {user_id}: {e}")
                failed += 1

        await update.message.reply_text(
            f"✅ پیام به {success} کاربر ارسال شد\n"
            f"❌ تعداد ناموفق: {failed}\n"
            f"🔹 کل کاربران: {len(verified_users)}"
        )
        context.user_data.pop('admin_broadcast_mode', None)
    else:
        context.user_data['admin_broadcast_mode'] = True
        await update.message.reply_text(
            "📢 پیام خود را برای کاربران تایید شده وارد کنید:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ لغو", callback_data="cancel_broadcast")]])
        )

async def list_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.from_user.id != ADMIN_CHAT_ID:
        await update.message.reply_text("❌ شما دسترسی ندارید!")
        return

    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute('SELECT * FROM verified_users')
    users = cur.fetchall()
    conn.close()

    if not users:
        await update.message.reply_text("هنوز کاربری تایید نشده است")
        return

    message = "📋 لیست کاربران تایید شده:\n\n"
    for user in users:
        message += (
            f"👤 نام: {user[3]}\n"
            f"📞 تلفن: {user[2]}\n"
            f"🆔 آیدی: {user[1]}\n"
            f"📌 کد ملی: {user[4]}\n"
            f"🗓 تاریخ ثبت: {user[5]}\n"
            "──────────────────\n"
        )

    await update.message.reply_text(message)

async def remove_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.from_user.id != ADMIN_CHAT_ID:
        await update.message.reply_text("❌ شما دسترسی ندارید!")
        return

    if not context.args:
        await update.message.reply_text("⚠️ لطفاً آیدی کاربر را وارد کنید:\n/remove_user <آیدی>")
        return

    user_id = context.args[0]

    try:
        conn = sqlite3.connect(DB_NAME)
        cur = conn.cursor()
        cur.execute('SELECT * FROM verified_users WHERE user_id = ?', (user_id,))
        user = cur.fetchone()

        if user:
            cur.execute('DELETE FROM verified_users WHERE user_id = ?', (user_id,))
            conn.commit()
            await update.message.reply_text(
                f"✅ کاربر با مشخصات زیر حذف شد:\n\n"
                f"👤 نام: {user[3]}\n"
                f"📞 تلفن: {user[2]}\n"
                f"🆔 آیدی: {user_id}\n"
                f"📌 کد ملی: {user[4]}"
            )
        else:
            await update.message.reply_text(f"❌ کاربر با آیدی {user_id} یافت نشد.")
    except Exception as e:
        logger.error(f"Error removing user: {e}")
        await update.message.reply_text("❌ خطا در حذف کاربر")
    finally:
        conn.close()

async def export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.from_user.id != ADMIN_CHAT_ID:
        await update.message.reply_text("❌ شما دسترسی ندارید!")
        return

    try:
        excel_file = create_excel_file()

        if not excel_file:
            await update.message.reply_text("❌ هیچ کاربری برای خروجی وجود ندارد")
            return

        await context.bot.send_chat_action(
            chat_id=update.message.chat_id, 
            action="upload_document"
        )

        await update.message.reply_document(
            document=open(excel_file, 'rb'),
            filename='users.xlsx',
            caption='📊 لیست کامل کاربران تایید شده'
        )

        os.remove(excel_file)

    except Exception as e:
        logger.error(f"Error in Excel export: {e}")
        await update.message.reply_text("❌ خطا در تولید فایل اکسل")

async def send_signal(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.from_user.id != ADMIN_CHAT_ID:
        await update.message.reply_text("❌ شما دسترسی ندارید!")
        return

    if context.user_data.get('awaiting_signal'):
        try:
            args = context.args
            if len(args) < 4:
                raise ValueError("Invalid arguments")

            entry = args[0]
            sl = args[1]
            tp = args[2]
            leverage = args[3]

            success, failed, total = await send_signal_to_users(context, entry, sl, tp, leverage)

            # تشخیص نوع پوزیشن برای گزارش ادمین
            position_type = "Long" if float(sl) < float(entry) else "Short"
            loss_percent = abs((float(sl) - float(entry)) / float(entry)) * 100

            await update.message.reply_text(
                f"✅ سیگنال با موفقیت ارسال شد!\n\n"
                f"🔹 تعداد موفق: {success}\n"
                f"🔹 تعداد ناموفق: {failed}\n"
                f"🔹 کل کاربران: {total}\n\n"
                f"📊 اطلاعات سیگنال:\n"
                f"📍 نوع پوزیشن: {position_type}\n"
                f"🎯 ورود: {entry}\n"
                f"🛑 SL: {sl}\n"
                f"✅ TP: {tp}\n"
                f"⚖️ اهرم: {leverage}x\n"
                f"📉 درصد ضرر: {loss_percent:.2f}%"
            )

            context.user_data.pop('awaiting_signal', None)
        except Exception as e:
            logger.error(f"Error sending signal: {e}")
            await update.message.reply_text(
                "❌ فرمت صحیح: /send_signal <entry> <sl> <tp> <leverage>\n"
                "مثال: /send_signal 50000 49500 52000 10"
            )
    else:
        context.user_data['awaiting_signal'] = True
        await update.message.reply_text(
            "📈 لطفاً اطلاعات سیگنال را به فرمت زیر وارد کنید:\n\n"
            "<code>/send_signal [قیمت ورود] [حد ضرر] [حد سود] [اهرم]</code>\n\n"
            "مثال:\n"
            "<code>/send_signal 50000 49500 52000 10</code>",
            parse_mode="HTML"
        )

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "register_signal":
        context.user_data.clear()
        context.user_data['registration_step'] = 'awaiting_phone'
        await query.edit_message_text("لطفاً شماره خود را با استفاده از دکمه زیر ارسال کنید:")
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="روی دکمه زیر بزنید:",
            reply_markup=phone_request_keyboard()
        )

    elif query.data == "support":
        await query.edit_message_text(
            "برای پشتیبانی با @ZerzangSupport تماس بگیرید.",
            reply_markup=back_button()
        )

    elif query.data == "calc_stop_loss":
        context.user_data.clear()
        context.user_data['awaiting_stop_loss'] = True
        await query.edit_message_text(
            "لطفاً مقادیر را به صورت زیر وارد کنید:\n\n"
            "🔹 بالایی: سرمایه دلاری شما\n"
            "🔸 پایینی: میزان ضرر دلاری مورد نظر\n\n"
            "مثال:\n"
            "20\n"
            "3",
            reply_markup=back_button()
        )

    elif query.data == "exchanges_brokers":
        msg = (
            "🔗 صرافی‌ها و بروکرهای ویژه:\n\n"
            "1. [Binance](https://www.binance.com/)\n"
            "2. [Bybit](https://www.bybit.com/)\n"
            "3. [Deribit](https://www.deribit.com/)\n"
            "4. [OKX](https://www.okx.com/)\n"
            "5. [MEXC](https://www.mexc.com/)"
        )
        await query.edit_message_text(msg, parse_mode="Markdown", reply_markup=back_button())

    elif query.data == "back_to_menu":
        context.user_data.clear()
        await query.edit_message_text(
            "لطفا گزینه مورد نظر را انتخاب کنید:",
            reply_markup=start_keyboard()
        )

    elif query.data == "cancel_broadcast":
        context.user_data.pop('admin_broadcast_mode', None)
        await query.edit_message_text("❌ ارسال پیام لغو شد.")

    elif query.data.startswith(("verify_payment:", "reject_payment:")):
        action, user_id = query.data.split(":")[0], int(query.data.split(":")[1])
        pending = get_pending_verification(user_id)

        if not pending:
            await query.edit_message_text("❌ درخواست تأیید یافت نشد!")
            return

        if "verify" in action:
            user_data = {
                'user_id': pending[1],
                'phone': pending[2],
                'full_name': pending[3],
                'nid': pending[4]
            }
            save_verified_user(user_data)
            remove_pending_verification(user_id)

            await context.bot.send_message(
                chat_id=user_id,
                text="✅ پرداخت شما تایید شد! ثبت‌نام تکمیل گردید."
            )
            await query.edit_message_text(
                f"✅ کاربر جدید ثبت شد:\n\n"
                f"👤 نام: {user_data['full_name']}\n"
                f"📞 تلفن: {user_data['phone']}\n"
                f"🆔 آیدی: {user_id}\n"
                f"📌 کد ملی: {user_data['nid']}\n"
                f"🗓 تاریخ ثبت: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                reply_markup=None
            )
        else:
            remove_pending_verification(user_id)
            await context.bot.send_message(
                chat_id=user_id,
                text="❌ پرداخت شما رد شد. لطفاً با پشتیبانی تماس بگیرید."
            )
            await query.edit_message_text(
                f"❌ پرداخت کاربر رد شد.\n\n"
                f"توسط ادمین در تاریخ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} انجام شد",
                reply_markup=None
            )
    
    elif query.data.startswith("capital_percent:"):
        parts = query.data.split(":")
        percent = float(parts[1])
        leverage = float(parts[2])
        
        context.user_data['capital_percent'] = percent
        context.user_data['leverage'] = leverage
        context.user_data['awaiting_capital'] = True
        
        await query.edit_message_text(
            f"📊 شما درصد <b>{percent}%</b> سرمایه را انتخاب کردید.\n\n"
            "💰 لطفاً میزان سرمایه دلاری خود را وارد کنید:\n"
            "مثال: 1000",
            parse_mode="HTML",
            reply_markup=back_button()
        )

async def message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message:
        return

    if update.message.from_user.id == ADMIN_CHAT_ID and context.user_data.get('admin_broadcast_mode'):
        await admin_broadcast(update, context)
        return

    text = update.message.text.strip() if update.message.text else ""

    if text == "🔙 بازگشت به منو":
        context.user_data.clear()
        await update.message.reply_text(
            "لطفا گزینه مورد نظر را انتخاب کنید:",
            reply_markup=start_keyboard()
        )
        return

    if context.user_data.get('awaiting_stop_loss'):
        try:
            parts = [p.strip() for p in text.split("\n") if p.strip()]
            if len(parts) < 2:
                raise ValueError("Invalid input")
                
            capital = float(parts[0])
            loss = float(parts[1])
            percent = (loss / capital) * 100
            msg = (
                f"📊 محاسبه حد ضرر:\n\n"
                f"💰 سرمایه شما: {capital} دلار\n"
                f"📉 ضرر مورد نظر: {loss} دلار\n"
                f"📈 درصد ضرر: {percent:.2f}%"
            )
            context.user_data.clear()
        except Exception as e:
            logger.error(f"Error in stop loss calculation: {e}")
            msg = "❌ لطفاً مقادیر را دقیقاً به این فرمت وارد کنید:\n20\n3"
        await update.message.reply_text(msg, reply_markup=back_button())
    
    elif context.user_data.get('awaiting_capital'):
        try:
            capital = float(text)
            percent = context.user_data['capital_percent']
            leverage = context.user_data['leverage']
            
            trade_amount = calculate_trade_amount(capital, percent, leverage)
            
            # دریافت اطلاعات سیگنال از پیام اصلی
            original_message = context.user_data.get('signal_message')
            if original_message:
                # استخراج entry و sl از پیام سیگنال
                entry = None
                sl = None
                for line in original_message.split('\n'):
                    if 'قیمت ورود' in line:
                        entry = line.split('<b>')[1].split('</b>')[0].strip()
                    elif 'حد ضرر' in line:
                        sl = line.split('<b>')[1].split('</b>')[0].strip()
                
                if entry and sl:
                    loss_amount = calculate_loss_amount(trade_amount, entry, sl)
                else:
                    loss_amount = 0
            else:
                loss_amount = 0
            
            msg = (
                f"📊 محاسبات سرمایه‌گذاری:\n\n"
                f"💵 سرمایه شما: <b>{capital:.2f} دلار</b>\n"
                f"📊 درصد انتخاب شده: <b>{percent}%</b>\n"
                f"⚖️ اهرم: <b>{leverage}x</b>\n\n"
                f"💳 میزان سرمایه ورودی به معامله: <b>{trade_amount:.2f} دلار</b>\n"
                f"📉 ضرر دلاری در صورت استاپ خوردن: <b>{loss_amount:.2f} دلار</b>"
            )
            
            context.user_data.clear()
            await update.message.reply_text(msg, parse_mode="HTML", reply_markup=back_button())
        except ValueError:
            await update.message.reply_text("❌ لطفاً یک عدد معتبر وارد کنید (مثال: 1000)")
        except Exception as e:
            logger.error(f"Error in capital calculation: {e}")
            await update.message.reply_text("❌ خطا در محاسبه سرمایه مورد نیاز")

    elif context.user_data.get('registration_step') == 'awaiting_name_nid':
        parts = text.split()
        if len(parts) >= 2:
            full_name = ' '.join(parts[:-1])
            nid = parts[-1]
            context.user_data['full_name'] = full_name
            context.user_data['nid'] = nid
            context.user_data['registration_step'] = 'awaiting_payment'
            await update.message.reply_text(
                "✅ اطلاعات شما ثبت شد.\nلطفاً عکس فیش واریزی را ارسال کنید:",
                reply_markup=back_button()
            )
        else:
            await update.message.reply_text("❌ فرمت صحیح: نام و نام خانوادگی + کد ملی\nمثال: علی رضایی 1234567890")

async def contact_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    contact = update.message.contact
    if context.user_data.get('registration_step') == 'awaiting_phone':
        context.user_data['phone_number'] = contact.phone_number
        context.user_data['user_id'] = contact.user_id
        context.user_data['registration_step'] = 'awaiting_name_nid'
        await update.message.reply_text(
            "✅ شماره شما ثبت شد.\n\n"
            "لطفاً نام و نام خانوادگی و کد ملی خود را به این صورت وارد کنید:\n\n"
            "مثال:\nعلی رضایی 1234567890",
            reply_markup=back_button()
        )

async def photo_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get('registration_step') == 'awaiting_payment':
        file_id = update.message.photo[-1].file_id

        user_data = {
            'user_id': context.user_data['user_id'],
            'phone': context.user_data['phone_number'],
            'full_name': context.user_data['full_name'],
            'nid': context.user_data['nid'],
            'file_id': file_id
        }
        save_pending_verification(user_data)

        await update.message.reply_text(
            "✅ فیش شما دریافت شد و در حال بررسی است.",
            reply_markup=start_keyboard()
        )
        
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ تایید پرداخت", callback_data=f"verify_payment:{context.user_data['user_id']}")],
            [InlineKeyboardButton("❌ رد پرداخت", callback_data=f"reject_payment:{context.user_data['user_id']}")]
        ])
        
        await context.bot.send_photo(
            chat_id=ADMIN_CHAT_ID,
            photo=file_id,
            caption=(
                "📌 درخواست ثبت‌نام جدید:\n\n"
                f"👤 نام: {context.user_data['full_name']}\n"
                f"🆔 کد ملی: {context.user_data['nid']}\n"
                f"📞 شماره: {context.user_data['phone_number']}\n"
                f"🆔 آیدی کاربر: {context.user_data['user_id']}"
            ),
            reply_markup=keyboard
        )

def main() -> None:
    start_keep_alive()

    web_thread = threading.Thread(target=run_web_server, daemon=True)
    web_thread.start()
    logger.info(f"🌐 سرور وب روی پورت {PORT} فعال شد")
    
    init_db()
    
    application = Application.builder().token(TOKEN).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("admin23", admin_broadcast))
    application.add_handler(CommandHandler("remove_user", remove_user))
    application.add_handler(CommandHandler("list_users", list_users))
    application.add_handler(CommandHandler("export", export_excel))
    application.add_handler(CommandHandler("send_signal", send_signal))
    application.add_handler(CallbackQueryHandler(button_handler))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, message_handler))
    application.add_handler(MessageHandler(filters.CONTACT, contact_handler))
    application.add_handler(MessageHandler(filters.PHOTO, photo_handler))

    logger.info("✅ ربات تلگرام در حال اجراست...")
    application.run_polling()

if __name__ == '__main__':
    main()